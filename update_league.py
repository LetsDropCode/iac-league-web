import pandas as pd
import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from update_engine import process_league

# -----------------------------------
# RUN ENGINE
# -----------------------------------
run_table, walk_table, _, _ = process_league()


# -----------------------------------
# CLEAN TABLE FOR EXPORT
# -----------------------------------
def clean_table(df):
    if df.empty:
        return df

    df = df.copy()

    # Remove backend columns
    df = df.drop(columns=["AthleteID"], errors="ignore")

    # Sort nicely
    if "Gender" in df.columns and "Rank" in df.columns:
        df = df.sort_values(["Gender", "Rank"])

    return df


run_table_clean = clean_table(run_table)
walk_table_clean = clean_table(walk_table)


# -----------------------------------
# CREATE EXCEL OUTPUT
# -----------------------------------
with pd.ExcelWriter("league_tables.xlsx", engine="openpyxl") as writer:

    def format_sheet(ws):
        ws.freeze_panes = "A2"

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto width
        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        # Center columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if ws.cell(row=1, column=cell.column).value in ["Rank", "Races Completed"]:
                    cell.alignment = Alignment(horizontal="center")


    # ---------------------------
    # RUNNERS
    # ---------------------------
    if not run_table_clean.empty:
        run_table_clean.to_excel(writer, sheet_name="Runners", index=False)
        format_sheet(writer.sheets["Runners"])


    # ---------------------------
    # WALKERS
    # ---------------------------
    if not walk_table_clean.empty:
        walk_table_clean.to_excel(writer, sheet_name="Walkers", index=False)
        format_sheet(writer.sheets["Walkers"])


    # ---------------------------
    # GENDER SPLIT (RUNNERS)
    # ---------------------------
    if not run_table_clean.empty and "Gender" in run_table_clean.columns:
        for gender in run_table_clean["Gender"].dropna().unique():

            subset = run_table_clean[run_table_clean["Gender"] == gender]

            if subset.empty:
                continue

            sheet_name = f"Run_{gender}"[:30]

            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            format_sheet(writer.sheets[sheet_name])


    # ---------------------------
    # CATEGORY SPLIT (RUNNERS)
    # 🔥 NEW
    # ---------------------------
    if not run_table_clean.empty and "PointsCategory" in run_table.columns:

        for cat in run_table["PointsCategory"].dropna().unique():

            subset = run_table[run_table["PointsCategory"] == cat]

            if subset.empty:
                continue

            subset = clean_table(subset)

            sheet_name = f"Cat_{cat}"[:30]

            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            format_sheet(writer.sheets[sheet_name])


    # ---------------------------
    # WALK CATEGORY SPLIT (OPTIONAL)
    # ---------------------------
    if not walk_table_clean.empty and "PointsCategory" in walk_table.columns:

        for cat in walk_table["PointsCategory"].dropna().unique():

            subset = walk_table[walk_table["PointsCategory"] == cat]

            if subset.empty:
                continue

            subset = clean_table(subset)

            sheet_name = f"Walk_{cat}"[:30]

            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            format_sheet(writer.sheets[sheet_name])


print("✅ Excel generated (clean + category split)")